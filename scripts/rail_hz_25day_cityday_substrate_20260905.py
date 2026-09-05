from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import json
import statistics
from collections import Counter, defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Any

TIME_FORMAT = "%Y-%m-%d %H:%M:%S"
MAX_VALID_TRIP_SEC = 8 * 3600
EXPECTED_DAYS = 25
EXPECTED_ROWS = 58_637_237
EXPECTED_VALID_TRIPS = 9_686_862


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def quantile(values: list[float], p: float) -> float | None:
    if not values:
        return None
    values = sorted(values)
    return values[round((len(values) - 1) * p)]


def sod(dt: datetime) -> int:
    return dt.hour * 3600 + dt.minute * 60 + dt.second


def bin_label(index: int, width_sec: int = 900) -> str:
    sec = index * width_sec
    h = (sec // 3600) % 24
    m = (sec % 3600) // 60
    return f"{h:02d}:{m:02d}"


def profile_day(input_path: Path, source_file: str, output_path: Path) -> dict[str, Any]:
    rows = 0
    tmin: datetime | None = None
    tmax: datetime | None = None
    status_counts: Counter[str] = Counter()
    line_counts: Counter[str] = Counter()
    pay_counts: Counter[str] = Counter()
    station_counts: Counter[str] = Counter()
    event_bins: Counter[int] = Counter()
    entry_bins: Counter[int] = Counter()
    exit_bins: Counter[int] = Counter()
    users: dict[str, list[Any]] = {}

    with gzip.open(input_path, "rt", encoding="utf-8-sig", errors="strict", newline="") as f:
        reader = csv.DictReader(f)
        columns = reader.fieldnames or []
        required = {"time", "lineID", "stationID", "status", "userID"}
        missing = required - set(columns)
        if missing:
            raise SystemExit(f"missing required columns: {sorted(missing)}")

        for row in reader:
            rows += 1
            try:
                t = datetime.strptime(row["time"], TIME_FORMAT)
            except Exception as exc:
                raise SystemExit(f"bad time at row {rows}: {exc}") from exc
            if tmin is None or t < tmin:
                tmin = t
            if tmax is None or t > tmax:
                tmax = t

            status = row["status"]
            line = row.get("lineID", "")
            station = row["stationID"]
            pay = row.get("payType", "")
            status_counts[status] += 1
            line_counts[line] += 1
            pay_counts[pay] += 1
            station_counts[station] += 1
            event_bins[sod(t) // 900] += 1
            if status == "1":
                entry_bins[sod(t) // 900] += 1
            elif status == "0":
                exit_bins[sod(t) // 900] += 1

            user = row["userID"]
            event = (t, status, line, station)
            rec = users.get(user)
            if rec is None:
                users[user] = [1, event, None]
            elif rec[0] == 1:
                rec[0] = 2
                rec[2] = event
            else:
                rec[0] += 1

    unique_user_count = len(users)
    pair_sequence: Counter[str] = Counter()
    valid_count = 0
    valid_same_line = 0
    valid_cross_line = 0
    travel_times: list[float] = []
    valid_entry_min: datetime | None = None
    valid_entry_max: datetime | None = None
    valid_exit_min: datetime | None = None
    valid_exit_max: datetime | None = None
    valid_origin_stations: Counter[str] = Counter()
    valid_destination_stations: Counter[str] = Counter()
    valid_od: Counter[str] = Counter()
    exact_two_users = 0

    for count, first, second in users.values():
        if count != 2 or second is None:
            continue
        exact_two_users += 1
        a, b = first, second
        if b[0] < a[0]:
            a, b = b, a
        pair_sequence[f"{a[1]}->{b[1]}"] += 1
        if a[1] != "1" or b[1] != "0":
            continue
        duration = (b[0] - a[0]).total_seconds()
        if not 0 <= duration <= MAX_VALID_TRIP_SEC:
            continue
        valid_count += 1
        travel_times.append(duration)
        valid_same_line += int(a[2] == b[2])
        valid_cross_line += int(a[2] != b[2])
        valid_origin_stations[a[3]] += 1
        valid_destination_stations[b[3]] += 1
        valid_od[f"{a[3]}->{b[3]}"] += 1
        if valid_entry_min is None or a[0] < valid_entry_min:
            valid_entry_min = a[0]
        if valid_entry_max is None or a[0] > valid_entry_max:
            valid_entry_max = a[0]
        if valid_exit_min is None or b[0] < valid_exit_min:
            valid_exit_min = b[0]
        if valid_exit_max is None or b[0] > valid_exit_max:
            valid_exit_max = b[0]

    del users

    if tmin is None or tmax is None:
        raise SystemExit("empty AFC file")

    source_date = source_file.removeprefix("record_").removesuffix(".csv.gz")
    observed_day_dates = sorted({tmin.date().isoformat(), tmax.date().isoformat()})
    support_start_sec = sod(valid_entry_min) if valid_entry_min else None
    support_end_sec = sod(valid_exit_max) if valid_exit_max else None

    result = {
        "schema": "rail.hz-cityday-substrate-day.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "source_file": source_file,
        "source_date": source_date,
        "source_sha256": sha256_file(input_path),
        "source_bytes": input_path.stat().st_size,
        "columns": columns,
        "rows": rows,
        "raw_time_min": tmin.isoformat(sep=" "),
        "raw_time_max": tmax.isoformat(sep=" "),
        "raw_dates_observed": observed_day_dates,
        "status_counts": dict(status_counts),
        "line_counts": dict(line_counts),
        "pay_type_counts": dict(pay_counts),
        "observed_station_count": len(station_counts),
        "observed_station_ids": sorted(station_counts),
        "unique_user_count": unique_user_count,
        "exact_two_event_users": exact_two_users,
        "exact_two_event_status_sequence": dict(pair_sequence),
        "valid_full_journey": {
            "definition": "exactly two events ordered 1->0 with duration in [0, 8h]",
            "count": valid_count,
            "same_line_count": valid_same_line,
            "cross_line_count": valid_cross_line,
            "cross_line_share": valid_cross_line / valid_count if valid_count else None,
            "travel_time_sec": {
                "median": quantile(travel_times, 0.5),
                "p90": quantile(travel_times, 0.9),
                "p99": quantile(travel_times, 0.99),
                "max": max(travel_times) if travel_times else None,
            },
            "first_valid_entry": valid_entry_min.isoformat(sep=" ") if valid_entry_min else None,
            "last_valid_entry": valid_entry_max.isoformat(sep=" ") if valid_entry_max else None,
            "first_valid_exit": valid_exit_min.isoformat(sep=" ") if valid_exit_min else None,
            "last_valid_exit": valid_exit_max.isoformat(sep=" ") if valid_exit_max else None,
            "observed_service_support_start_sec_of_day": support_start_sec,
            "observed_service_support_end_sec_of_day": support_end_sec,
            "origin_station_count": len(valid_origin_stations),
            "destination_station_count": len(valid_destination_stations),
            "od_pair_count": len(valid_od),
        },
        "activity_15min": {
            "all_events": {bin_label(i): event_bins.get(i, 0) for i in range(96)},
            "entry_events": {bin_label(i): entry_bins.get(i, 0) for i in range(96)},
            "exit_events": {bin_label(i): exit_bins.get(i, 0) for i in range(96)},
        },
        "evidence_manifest": {
            "afc_events": "OBSERVED_AFC",
            "entry_exit_semantics": "EMPIRICALLY_QUALIFIED_STATUS_ORDERING",
            "valid_journeys": "DETERMINISTIC_DERIVATION_FROM_OBSERVED_AFC",
            "service_events": "NOT_ASSIGNED_AT_SUBSTRATE_STAGE",
            "network_topology": "STRUCTURAL_PRIOR_SEPARATE_STATIC_MANIFEST",
            "auxiliary_timetable": "STRUCTURAL_PRIOR_ONLY_NOT_REALIZED_2019_TRUTH",
        },
        "privacy_rule": "userID/deviceID values are used transiently for pairing/counting and are never exported",
    }
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def components(adj: list[list[int]]) -> list[list[int]]:
    n = len(adj)
    seen: set[int] = set()
    out: list[list[int]] = []
    for start in range(n):
        if start in seen:
            continue
        q = deque([start])
        seen.add(start)
        comp: list[int] = []
        while q:
            i = q.popleft()
            comp.append(i)
            for j, value in enumerate(adj[i]):
                if value and j not in seen:
                    seen.add(j)
                    q.append(j)
        out.append(comp)
    return out


def median(values: list[float]) -> float | None:
    return statistics.median(values) if values else None


def profile_static(roadmap: Path, timetable: Path, output_path: Path) -> dict[str, Any]:
    with roadmap.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if len(rows) < 2 or len(rows[0]) < 2:
        raise SystemExit("invalid roadmap matrix")
    header = rows[0][1:]
    station_labels = [r[0] for r in rows[1:]]
    matrix: list[list[int]] = []
    for row in rows[1:]:
        vals = [int(float(v)) for v in row[1:]]
        matrix.append(vals)
    n = len(matrix)
    if any(len(r) != n for r in matrix):
        raise SystemExit(f"roadmap is not square: {n} rows")
    undirected_edges = sum(matrix[i][j] != 0 for i in range(n) for j in range(i + 1, n))
    comps = components(matrix)

    from openpyxl import load_workbook

    wb = load_workbook(timetable, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    columns = [str(v) if v is not None else "" for v in next(it)]
    idx = {name: i for i, name in enumerate(columns)}
    required = ["Train ID", "Line", "Direction", "Station Order", "Station Name", "Arrival Time", "Departure Time", "Dwell Time(s)"]
    missing = [x for x in required if x not in idx]
    if missing:
        raise SystemExit(f"timetable missing columns: {missing}")

    timetable_rows = 0
    trains: set[str] = set()
    lines: set[str] = set()
    directions: set[str] = set()
    absolute_dates: set[str] = set()
    absolute_min: datetime | None = None
    absolute_max: datetime | None = None
    by_train: dict[str, list[tuple[int, str, str, str, datetime | None, datetime | None, float | None]]] = defaultdict(list)
    dwell_groups: dict[tuple[str, str, str], list[float]] = defaultdict(list)

    for values in it:
        timetable_rows += 1
        train = str(values[idx["Train ID"]])
        line = str(values[idx["Line"]])
        direction = str(values[idx["Direction"]])
        order = int(values[idx["Station Order"]])
        station = str(values[idx["Station Name"]])
        arr = values[idx["Arrival Time"]]
        dep = values[idx["Departure Time"]]
        dwell_raw = values[idx["Dwell Time(s)"]]
        dwell = float(dwell_raw) if dwell_raw is not None else None
        arr = arr if isinstance(arr, datetime) else None
        dep = dep if isinstance(dep, datetime) else None
        trains.add(train)
        lines.add(line)
        directions.add(direction)
        for t in (arr, dep):
            if t is None:
                continue
            absolute_dates.add(t.date().isoformat())
            if absolute_min is None or t < absolute_min:
                absolute_min = t
            if absolute_max is None or t > absolute_max:
                absolute_max = t
        if dwell is not None and dwell >= 0:
            dwell_groups[(line, direction, station)].append(dwell)
        by_train[train].append((order, line, direction, station, arr, dep, dwell))
    wb.close()

    run_groups: dict[tuple[str, str, str, str], list[float]] = defaultdict(list)
    for stops in by_train.values():
        stops.sort(key=lambda x: x[0])
        for a, b in zip(stops, stops[1:]):
            if a[1] != b[1] or a[2] != b[2] or a[5] is None or b[4] is None:
                continue
            run = (b[4] - a[5]).total_seconds()
            if 0 < run <= 3600:
                run_groups[(a[1], a[2], a[3], b[3])].append(run)

    segment_priors = []
    for key, vals in sorted(run_groups.items()):
        line, direction, from_station, to_station = key
        segment_priors.append({
            "line": line,
            "direction": direction,
            "from_station": from_station,
            "to_station": to_station,
            "n": len(vals),
            "run_time_sec_median": median(vals),
            "run_time_sec_p10": quantile(vals, 0.1),
            "run_time_sec_p90": quantile(vals, 0.9),
        })
    dwell_priors = []
    for key, vals in sorted(dwell_groups.items()):
        line, direction, station = key
        dwell_priors.append({
            "line": line,
            "direction": direction,
            "station": station,
            "n": len(vals),
            "dwell_time_sec_median": median(vals),
            "dwell_time_sec_p10": quantile(vals, 0.1),
            "dwell_time_sec_p90": quantile(vals, 0.9),
        })

    result = {
        "schema": "rail.hz-cityday-static-substrate.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "network": {
            "source_file": roadmap.name,
            "sha256": sha256_file(roadmap),
            "matrix_node_count": n,
            "header_node_count": len(header),
            "station_label_count": len(station_labels),
            "undirected_edge_count": undirected_edges,
            "connected_component_count": len(comps),
            "component_sizes": sorted((len(c) for c in comps), reverse=True),
            "full_network_retained": True,
        },
        "auxiliary_timetable": {
            "source_file": timetable.name,
            "sha256": sha256_file(timetable),
            "rows": timetable_rows,
            "train_count": len(trains),
            "lines": sorted(lines),
            "directions": sorted(directions),
            "absolute_dates": sorted(absolute_dates),
            "absolute_time_min": absolute_min.isoformat(sep=" ") if absolute_min else None,
            "absolute_time_max": absolute_max.isoformat(sep=" ") if absolute_max else None,
            "evidence_class": "STRUCTURAL_SERVICE_PRIOR_ONLY",
            "absolute_timestamp_policy": "FORBIDDEN_AS_2019_REALIZED_TRUTH",
            "allowed_reuse": ["line_order", "direction", "station_sequence", "relative_segment_runtime", "dwell_organization", "service_organization_prior"],
        },
        "relative_service_structure_prior": {
            "segment_count": len(segment_priors),
            "segments": segment_priors,
            "dwell_context_count": len(dwell_priors),
            "dwell_contexts": dwell_priors,
        },
    }
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def aggregate(days_dir: Path, static_path: Path, output_path: Path) -> dict[str, Any]:
    day_paths = sorted(days_dir.glob("record_2019-01-*.cityday.json"))
    days = [json.loads(p.read_text(encoding="utf-8")) for p in day_paths]
    static = json.loads(static_path.read_text(encoding="utf-8"))
    expected_dates = [f"2019-01-{d:02d}" for d in range(1, 26)]
    actual_dates = [d["source_date"] for d in days]

    pooled_bins = Counter()
    pooled_entry_bins = Counter()
    pooled_exit_bins = Counter()
    for d in days:
        for label, n in d["activity_15min"]["all_events"].items():
            pooled_bins[label] += n
        for label, n in d["activity_15min"]["entry_events"].items():
            pooled_entry_bins[label] += n
        for label, n in d["activity_15min"]["exit_events"].items():
            pooled_exit_bins[label] += n

    all_labels = [bin_label(i) for i in range(96)]
    trough_label = min(all_labels, key=lambda x: pooled_bins[x]) if days else None
    total_rows = sum(d["rows"] for d in days)
    total_valid = sum(d["valid_full_journey"]["count"] for d in days)
    total_events = sum(pooled_bins.values())
    overnight_labels = [bin_label(i) for i in range(20)]
    overnight_events = sum(pooled_bins[x] for x in overnight_labels)

    gates = {
        "exactly_25_daily_sources": len(days) == EXPECTED_DAYS,
        "exact_date_range_2019_01_01_to_25": actual_dates == expected_dates,
        "prior_total_afc_rows_reproduced": total_rows == EXPECTED_ROWS,
        "prior_total_valid_full_journeys_reproduced": total_valid == EXPECTED_VALID_TRIPS,
        "every_day_has_valid_full_journeys": all(d["valid_full_journey"]["count"] > 0 for d in days),
        "static_full_network_node_count_81": static["network"]["matrix_node_count"] == 81,
        "static_network_single_component": static["network"]["connected_component_count"] == 1,
        "auxiliary_absolute_time_not_promoted_to_realized_truth": static["auxiliary_timetable"]["absolute_timestamp_policy"] == "FORBIDDEN_AS_2019_REALIZED_TRUTH",
    }
    qualified = all(gates.values())

    day_table = []
    for d in days:
        v = d["valid_full_journey"]
        day_table.append({
            "date": d["source_date"],
            "rows": d["rows"],
            "valid_full_journeys": v["count"],
            "observed_station_count": d["observed_station_count"],
            "first_valid_entry": v["first_valid_entry"],
            "last_valid_exit": v["last_valid_exit"],
            "od_pair_count": v["od_pair_count"],
            "cross_line_share": v["cross_line_share"],
            "travel_time_median_sec": v["travel_time_sec"]["median"],
            "source_sha256": d["source_sha256"],
        })

    result = {
        "schema": "rail.hz-25day-cityday-substrate-summary.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "status": "QUALIFIED_HZ25_CITY_DAY_SUBSTRATE" if qualified else "HZ25_CITY_DAY_SUBSTRATE_GATE_FAILED",
        "certification_unit": "CITY_DAY_FULL_NETWORK_FULL_SERVICE_DAY",
        "coverage": {
            "days": len(days),
            "dates": actual_dates,
            "total_afc_rows": total_rows,
            "total_valid_full_journeys": total_valid,
            "static_network_nodes": static["network"]["matrix_node_count"],
            "static_network_edges": static["network"]["undirected_edge_count"],
            "static_network_components": static["network"]["connected_component_count"],
        },
        "service_day_support": {
            "boundary_definition": "per city-day observed support is the first valid entry through the last valid exit; no peak-window cut is imposed",
            "pooled_15min_activity_trough": trough_label,
            "pooled_15min_activity_trough_events": pooled_bins[trough_label] if trough_label else None,
            "pooled_0000_0500_event_share": overnight_events / total_events if total_events else None,
            "cross_midnight_policy": "overnight activity is reported explicitly; any subsequent business-day reassignment must preserve all events and may not truncate the operating day",
        },
        "integrity_gates": gates,
        "evidence_classes": {
            "daily_afc": "OBSERVED_AFC",
            "valid_journeys": "DERIVED_FROM_OBSERVED_AFC",
            "network": "STRUCTURAL_PRIOR",
            "auxiliary_timetable": "STRUCTURAL_SERVICE_PRIOR_ONLY",
            "service_state": "NOT_YET_INFERRED",
        },
        "scope_invariants": {
            "peak_window_filter": False,
            "line_filter": False,
            "transfer_count_filter": False,
            "passenger_subsample": False,
            "full_network_static_state_retained": True,
        },
        "days": day_table,
        "pooled_activity_15min": {
            "all_events": {x: pooled_bins[x] for x in all_labels},
            "entry_events": {x: pooled_entry_bins[x] for x in all_labels},
            "exit_events": {x: pooled_exit_bins[x] for x in all_labels},
        },
        "privacy_rule": "no raw userID/deviceID values are persisted in derived substrate outputs",
        "next_stage_if_qualified": "DAY_SPECIFIC_SERVICE_INITIALIZATION",
    }
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": result["status"],
        "days": len(days),
        "total_afc_rows": total_rows,
        "total_valid_full_journeys": total_valid,
        "network_nodes": static["network"]["matrix_node_count"],
        "pooled_15min_activity_trough": trough_label,
        "pooled_0000_0500_event_share": result["service_day_support"]["pooled_0000_0500_event_share"],
        "integrity_gates": gates,
    }, ensure_ascii=False, indent=2))
    if not qualified:
        raise SystemExit(2)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)

    p_day = sub.add_parser("day")
    p_day.add_argument("--input", type=Path, required=True)
    p_day.add_argument("--source-file", required=True)
    p_day.add_argument("--output", type=Path, required=True)

    p_static = sub.add_parser("static")
    p_static.add_argument("--roadmap", type=Path, required=True)
    p_static.add_argument("--timetable", type=Path, required=True)
    p_static.add_argument("--output", type=Path, required=True)

    p_agg = sub.add_parser("aggregate")
    p_agg.add_argument("--days-dir", type=Path, required=True)
    p_agg.add_argument("--static", type=Path, required=True)
    p_agg.add_argument("--output", type=Path, required=True)

    args = parser.parse_args()
    if args.command == "day":
        profile_day(args.input, args.source_file, args.output)
    elif args.command == "static":
        profile_static(args.roadmap, args.timetable, args.output)
    elif args.command == "aggregate":
        aggregate(args.days_dir, args.static, args.output)


if __name__ == "__main__":
    main()

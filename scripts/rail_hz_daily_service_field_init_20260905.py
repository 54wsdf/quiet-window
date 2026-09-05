from __future__ import annotations

import argparse
import csv
import gzip
import json
import math
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np

BIN_S = 5
DAY_S = 24 * 3600
NBINS = DAY_S // BIN_S
CONTEXT_S = 2 * 3600
LAG_SEARCH_HALF_WIDTH_S = 120
MAX_EDGE_LAG_S = 1200

LINE_PATHS: dict[str, dict[str, Any]] = {
    "B_main": {"afc_line": "B", "aux_line": "L1", "nodes": list(range(0, 28))},
    "B_branch": {"afc_line": "B", "aux_line": "L1", "nodes": list(range(0, 21)) + list(range(28, 34))},
    "C_main": {"afc_line": "C", "aux_line": "L2", "nodes": list(range(34, 67))},
    "A_main": {"afc_line": "A", "aux_line": "L4", "nodes": [67, 68, 69, 70, 71, 72, 73, 74, 5, 75, 76, 77, 46, 78, 79, 80, 15, 16]},
}


def quantile(values: list[float], p: float) -> float | None:
    if not values:
        return None
    x = sorted(values)
    return float(x[round((len(x) - 1) * p)])


def moving_average(x: np.ndarray, width: int) -> np.ndarray:
    width = max(1, int(width))
    return np.convolve(np.asarray(x, dtype=float), np.ones(width, dtype=float) / width, mode="same")


def detect_events(counts: np.ndarray) -> list[dict[str, float]]:
    x = np.asarray(counts, dtype=float)
    short_bins = max(1, round(15 / BIN_S))
    background_bins = max(3, round(300 / BIN_S))
    short = moving_average(x, short_bins)
    background = moving_average(x, background_bins)
    residual = short - background
    score = residual / np.sqrt(np.maximum(background, 0.25))
    threshold = max(float(np.quantile(score, 0.90)), 0.50)
    left = np.r_[score[0], score[:-1]]
    right = np.r_[score[1:], score[-1]]
    edge_guard = max(1, background_bins // 2)
    interior = np.zeros(len(x), dtype=bool)
    if len(x) > 2 * edge_guard:
        interior[edge_guard:-edge_guard] = True
    candidates = np.flatnonzero((score >= threshold) & (score >= left) & (score >= right) & (residual > 0) & interior)
    ranked = candidates[np.argsort(-score[candidates], kind="stable")]
    selected: list[int] = []
    min_sep_bins = max(1, math.ceil(90 / BIN_S))
    for idx in ranked:
        i = int(idx)
        if all(abs(i - j) >= min_sep_bins for j in selected):
            selected.append(i)
    selected.sort()

    half = max(1, math.ceil(60 / BIN_S))
    out: list[dict[str, float]] = []
    for i in selected:
        lo = i
        hi = i
        while lo > 0 and i - lo < half and residual[lo - 1] > 0:
            lo -= 1
        while hi + 1 < len(x) and hi - i < half and residual[hi + 1] > 0:
            hi += 1
        raw = float(np.sum(x[lo : hi + 1]))
        bg = float(np.sum(background[lo : hi + 1]))
        out.append({
            "center_s": float((i + 0.5) * BIN_S),
            "start_s": float(lo * BIN_S),
            "end_s": float((hi + 1) * BIN_S),
            "mass": raw,
            "background_mass": bg,
            "excess_mass": max(0.0, raw - bg),
            "score": float(score[i]),
            "threshold": threshold,
        })
    return out


def highpass(x: np.ndarray) -> np.ndarray:
    return moving_average(x, 3) - moving_average(x, 60)


def corr_at_lag(a: np.ndarray, b: np.ndarray, lag_bins: int) -> float:
    if lag_bins > 0:
        x, y = a[:-lag_bins], b[lag_bins:]
    elif lag_bins < 0:
        x, y = a[-lag_bins:], b[:lag_bins]
    else:
        x, y = a, b
    if len(x) < 60 or x.std() <= 1e-12 or y.std() <= 1e-12:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


def best_lag(a: np.ndarray, b: np.ndarray, expected_s: float) -> tuple[float | None, float | None]:
    lo = max(BIN_S, int(math.floor((expected_s - LAG_SEARCH_HALF_WIDTH_S) / BIN_S)) * BIN_S)
    hi = min(MAX_EDGE_LAG_S, int(math.ceil((expected_s + LAG_SEARCH_HALF_WIDTH_S) / BIN_S)) * BIN_S)
    vals: list[tuple[float, int]] = []
    for lag_s in range(lo, hi + 1, BIN_S):
        c = corr_at_lag(a, b, lag_s // BIN_S)
        if np.isfinite(c):
            vals.append((c, lag_s))
    if not vals:
        return None, None
    c, lag_s = max(vals, key=lambda x: (x[0], -abs(x[1] - expected_s)))
    return float(lag_s), float(c)


def parse_timetable_priors(timetable: Path, output: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(timetable, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    columns = [str(x) if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(columns)}
    required = ["Train ID", "Line", "Direction", "Station Order", "Station Name", "Arrival Time", "Departure Time"]
    missing = [x for x in required if x not in idx]
    if missing:
        raise SystemExit(f"missing timetable columns: {missing}")

    trains: dict[tuple[str, str, str], list[tuple[int, str, datetime | None, datetime | None]]] = defaultdict(list)
    absolute_dates: set[str] = set()
    for row in rows:
        train = str(row[idx["Train ID"]])
        line = str(row[idx["Line"]])
        direction = str(row[idx["Direction"]])
        order = int(row[idx["Station Order"]])
        station = str(row[idx["Station Name"]])
        arr = row[idx["Arrival Time"]]
        dep = row[idx["Departure Time"]]
        arr = arr if isinstance(arr, datetime) else None
        dep = dep if isinstance(dep, datetime) else None
        for t in (arr, dep):
            if t is not None:
                absolute_dates.add(t.date().isoformat())
        trains[(line, direction, train)].append((order, station, arr, dep))
    wb.close()

    edge_values: dict[tuple[str, str, int, int], list[float]] = defaultdict(list)
    path_sequence_counts: Counter[str] = Counter()
    for path_id, meta in LINE_PATHS.items():
        aux_line = meta["aux_line"]
        base_nodes = list(meta["nodes"])
        target_len = len(base_nodes)
        for direction in ("Down", "Up"):
            numeric = base_nodes if direction == "Down" else list(reversed(base_nodes))
            for (line, d, train), stops in trains.items():
                if line != aux_line or d != direction:
                    continue
                stops = sorted(stops, key=lambda x: x[0])
                if len(stops) != target_len:
                    continue
                path_sequence_counts[f"{path_id}:{direction}"] += 1
                for a, b, u, v in zip(stops, stops[1:], numeric, numeric[1:]):
                    if a[2] is None or b[2] is None:
                        continue
                    lag = (b[2] - a[2]).total_seconds()
                    if 0 < lag <= MAX_EDGE_LAG_S:
                        edge_values[(path_id, direction, u, v)].append(float(lag))

    priors = []
    for path_id, meta in LINE_PATHS.items():
        base_nodes = list(meta["nodes"])
        for direction in ("Down", "Up"):
            numeric = base_nodes if direction == "Down" else list(reversed(base_nodes))
            for u, v in zip(numeric, numeric[1:]):
                vals = edge_values.get((path_id, direction, u, v), [])
                priors.append({
                    "path_id": path_id,
                    "afc_line": meta["afc_line"],
                    "aux_line": meta["aux_line"],
                    "direction": direction,
                    "from_station": u,
                    "to_station": v,
                    "n": len(vals),
                    "arrival_to_arrival_lag_sec_median": statistics.median(vals) if vals else None,
                    "arrival_to_arrival_lag_sec_p10": quantile(vals, 0.1),
                    "arrival_to_arrival_lag_sec_p90": quantile(vals, 0.9),
                })

    result = {
        "schema": "rail.hz-relative-service-prior.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "line_paths": LINE_PATHS,
        "source_absolute_dates": sorted(absolute_dates),
        "absolute_timestamp_policy": "FORBIDDEN_AS_2019_REALIZED_TRUTH",
        "prior_semantics": "only relative arrival-to-arrival propagation and route organization are reused",
        "path_sequence_counts": dict(path_sequence_counts),
        "edges": priors,
    }
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def load_exit_counts(input_path: Path) -> tuple[np.ndarray, dict[str, Any]]:
    counts = np.zeros((81, NBINS), dtype=np.uint32)
    rows = 0
    exit_rows = 0
    station_rows: Counter[int] = Counter()
    line_rows: Counter[str] = Counter()
    with gzip.open(input_path, "rt", encoding="utf-8-sig", errors="strict", newline="") as f:
        reader = csv.DictReader(f)
        required = {"time", "stationID", "status"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise SystemExit(f"missing AFC columns: {sorted(missing)}")
        for row in reader:
            rows += 1
            if row["status"] != "0":
                continue
            t = datetime.strptime(row["time"], "%Y-%m-%d %H:%M:%S")
            try:
                station = int(row["stationID"])
            except ValueError:
                continue
            if not 0 <= station < 81:
                continue
            k = (t.hour * 3600 + t.minute * 60 + t.second) // BIN_S
            if 0 <= k < NBINS:
                counts[station, k] += 1
                exit_rows += 1
                station_rows[station] += 1
                line_rows[row.get("lineID", "")] += 1
    meta = {
        "rows": rows,
        "exit_rows": exit_rows,
        "exit_station_count": len(station_rows),
        "exit_station_rows": {str(k): v for k, v in sorted(station_rows.items())},
        "exit_line_rows": dict(line_rows),
    }
    return counts, meta


def service_day(input_path: Path, source_file: str, prior_path: Path, substrate_path: Path, npz_output: Path, json_output: Path) -> dict[str, Any]:
    prior = json.loads(prior_path.read_text(encoding="utf-8"))
    substrate = json.loads(substrate_path.read_text(encoding="utf-8"))
    counts, meta = load_exit_counts(input_path)

    event_station: list[int] = []
    event_center: list[float] = []
    event_start: list[float] = []
    event_end: list[float] = []
    event_mass: list[float] = []
    event_excess: list[float] = []
    event_score: list[float] = []
    event_threshold: list[float] = []
    station_event_counts: dict[str, int] = {}

    for station in range(81):
        events = detect_events(counts[station]) if counts[station].sum() > 0 else []
        station_event_counts[str(station)] = len(events)
        for e in events:
            event_station.append(station)
            event_center.append(e["center_s"])
            event_start.append(e["start_s"])
            event_end.append(e["end_s"])
            event_mass.append(e["mass"])
            event_excess.append(e["excess_mass"])
            event_score.append(e["score"])
            event_threshold.append(e["threshold"])

    edge_contexts = []
    finite_afc_lag_contexts = 0
    total_prior_contexts = 0
    for edge in prior["edges"]:
        expected = edge["arrival_to_arrival_lag_sec_median"]
        u = int(edge["from_station"])
        v = int(edge["to_station"])
        if expected is None:
            continue
        for context_start in range(0, DAY_S, CONTEXT_S):
            context_end = min(DAY_S, context_start + CONTEXT_S)
            a = context_start // BIN_S
            b = context_end // BIN_S
            x = counts[u, a:b].astype(float)
            y = counts[v, a:b].astype(float)
            x_mass = float(x.sum())
            y_mass = float(y.sum())
            total_prior_contexts += 1
            if x_mass >= 20 and y_mass >= 20:
                lag, corr = best_lag(highpass(x), highpass(y), float(expected))
            else:
                lag, corr = None, None
            if lag is not None:
                finite_afc_lag_contexts += 1
            edge_contexts.append({
                "path_id": edge["path_id"],
                "direction": edge["direction"],
                "from_station": u,
                "to_station": v,
                "context_start_s": context_start,
                "context_end_s": context_end,
                "upstream_exit_mass": x_mass,
                "downstream_exit_mass": y_mass,
                "structural_prior_lag_s": expected,
                "afc_supported_lag_s": lag,
                "afc_correlation": corr,
                "afc_minus_structural_lag_s": lag - float(expected) if lag is not None else None,
                "effective_initial_lag_s": lag if lag is not None else expected,
                "evidence_class": "AFC_PLUS_STRUCTURAL_PRIOR" if lag is not None else "STRUCTURAL_PRIOR_FALLBACK",
            })

    np.savez_compressed(
        npz_output,
        bin_s=np.array([BIN_S], dtype=np.int32),
        exit_counts=counts,
        event_station=np.asarray(event_station, dtype=np.int16),
        event_center_s=np.asarray(event_center, dtype=np.float32),
        event_start_s=np.asarray(event_start, dtype=np.float32),
        event_end_s=np.asarray(event_end, dtype=np.float32),
        event_mass=np.asarray(event_mass, dtype=np.float32),
        event_excess_mass=np.asarray(event_excess, dtype=np.float32),
        event_score=np.asarray(event_score, dtype=np.float32),
        event_threshold=np.asarray(event_threshold, dtype=np.float32),
    )

    no_event_nodes = [i for i in range(81) if station_event_counts[str(i)] == 0]
    source_date = source_file.removeprefix("record_").removesuffix(".csv.gz")
    result = {
        "schema": "rail.hz-day-specific-service-field-init.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "source_date": source_date,
        "source_file": source_file,
        "state_representation": "FACTORIZED_PASSENGER_FACING_EVENT_AND_PROPAGATION_FIELD",
        "state_semantics": "initial latent service field S_d^(0), not observed ATS and not exact train identity",
        "time_scope": "FULL_CALENDAR_SOURCE_DAY_WITH_FULL_SERVICE_SUPPORT_RETAINED",
        "substrate_service_support": substrate.get("valid_full_journey", {}),
        "afc_profile": meta,
        "event_field": {
            "bin_s": BIN_S,
            "event_count": len(event_station),
            "station_event_counts": station_event_counts,
            "stations_with_events": sum(v > 0 for v in station_event_counts.values()),
            "nodes_without_detected_exit_events": no_event_nodes,
            "evidence_class": "AFC_INFERRED_PASSENGER_FACING_EVENT",
            "event_time_is_train_arrival_truth": False,
        },
        "propagation_field": {
            "context_width_s": CONTEXT_S,
            "lag_search_half_width_s": LAG_SEARCH_HALF_WIDTH_S,
            "total_structural_prior_contexts": total_prior_contexts,
            "afc_supported_lag_contexts": finite_afc_lag_contexts,
            "afc_supported_share": finite_afc_lag_contexts / total_prior_contexts if total_prior_contexts else None,
            "contexts": edge_contexts,
        },
        "uncertainty_policy": {
            "local_event_time": "pulse width and AFC detector score retained",
            "edge_lag": "AFC-supported lag used when estimable; otherwise structural prior remains explicit fallback",
            "unresolved_nodes": "retained as unresolved latent service support rather than deleted",
            "direction": "encoded by structural path orientation; local station events remain potentially direction-ambiguous until joint inference",
        },
        "scope_invariants": {
            "peak_window_filter": False,
            "line_filter": False,
            "passenger_subsample": False,
            "absolute_auxiliary_timetable_timestamp_used": False,
        },
        "machine_state_file": npz_output.name,
    }
    json_output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def aggregate(days_dir: Path, prior_path: Path, output: Path) -> dict[str, Any]:
    files = sorted(days_dir.glob("record_2019-01-*.service_init.json"))
    days = [json.loads(p.read_text(encoding="utf-8")) for p in files]
    prior = json.loads(prior_path.read_text(encoding="utf-8"))
    expected_dates = [f"2019-01-{d:02d}" for d in range(1, 26)]
    actual_dates = [d["source_date"] for d in days]
    prior_edges = [e for e in prior["edges"] if e["arrival_to_arrival_lag_sec_median"] is not None]
    gates = {
        "all_25_days_present": len(days) == 25 and actual_dates == expected_dates,
        "relative_prior_has_supported_edges": len(prior_edges) > 0,
        "all_days_have_passenger_facing_events": all(d["event_field"]["event_count"] > 0 for d in days),
        "all_days_preserve_full_scope": all(not d["scope_invariants"]["peak_window_filter"] and not d["scope_invariants"]["line_filter"] and not d["scope_invariants"]["passenger_subsample"] for d in days),
        "no_day_uses_absolute_auxiliary_timestamp": all(not d["scope_invariants"]["absolute_auxiliary_timetable_timestamp_used"] for d in days),
        "auxiliary_prior_policy_safe": prior["absolute_timestamp_policy"] == "FORBIDDEN_AS_2019_REALIZED_TRUTH",
    }
    qualified = all(gates.values())
    event_counts = [d["event_field"]["event_count"] for d in days]
    lag_shares = [d["propagation_field"]["afc_supported_share"] for d in days if d["propagation_field"]["afc_supported_share"] is not None]
    result = {
        "schema": "rail.hz-25day-service-field-init-summary.v1",
        "dataset_id": "CN_HZ_Tianchi_2019",
        "status": "QUALIFIED_HZ25_DAY_SPECIFIC_SERVICE_INIT" if qualified else "HZ25_SERVICE_INIT_GATE_FAILED",
        "state_representation": "FACTORIZED_PASSENGER_FACING_EVENT_AND_PROPAGATION_FIELD",
        "days": len(days),
        "dates": actual_dates,
        "integrity_gates": gates,
        "relative_prior": {
            "supported_directed_edges": len(prior_edges),
            "path_sequence_counts": prior["path_sequence_counts"],
        },
        "event_field": {
            "total_events": sum(event_counts),
            "daily_event_count_median": statistics.median(event_counts) if event_counts else None,
            "daily_event_count_min": min(event_counts) if event_counts else None,
            "daily_event_count_max": max(event_counts) if event_counts else None,
        },
        "propagation_field": {
            "daily_afc_supported_context_share_median": statistics.median(lag_shares) if lag_shares else None,
            "daily_afc_supported_context_share_min": min(lag_shares) if lag_shares else None,
            "daily_afc_supported_context_share_max": max(lag_shares) if lag_shares else None,
        },
        "scientific_boundary": {
            "observed_ats": False,
            "exact_train_identity": False,
            "day_specific_service_field": True,
            "uncertainty_retained": True,
            "next_update": "full-day bidirectional passenger-service R1B update",
        },
        "day_summaries": [{
            "date": d["source_date"],
            "events": d["event_field"]["event_count"],
            "stations_with_events": d["event_field"]["stations_with_events"],
            "afc_supported_lag_context_share": d["propagation_field"]["afc_supported_share"],
            "unresolved_nodes": d["event_field"]["nodes_without_detected_exit_events"],
        } for d in days],
    }
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": result["status"],
        "days": len(days),
        "total_events": result["event_field"]["total_events"],
        "median_afc_supported_context_share": result["propagation_field"]["daily_afc_supported_context_share_median"],
        "integrity_gates": gates,
    }, ensure_ascii=False, indent=2))
    if not qualified:
        raise SystemExit(2)
    return result


def main() -> None:
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="command", required=True)

    s = sub.add_parser("prior")
    s.add_argument("--timetable", type=Path, required=True)
    s.add_argument("--output", type=Path, required=True)

    s = sub.add_parser("day")
    s.add_argument("--input", type=Path, required=True)
    s.add_argument("--source-file", required=True)
    s.add_argument("--prior", type=Path, required=True)
    s.add_argument("--substrate", type=Path, required=True)
    s.add_argument("--npz-output", type=Path, required=True)
    s.add_argument("--json-output", type=Path, required=True)

    s = sub.add_parser("aggregate")
    s.add_argument("--days-dir", type=Path, required=True)
    s.add_argument("--prior", type=Path, required=True)
    s.add_argument("--output", type=Path, required=True)

    args = p.parse_args()
    if args.command == "prior":
        parse_timetable_priors(args.timetable, args.output)
    elif args.command == "day":
        service_day(args.input, args.source_file, args.prior, args.substrate, args.npz_output, args.json_output)
    elif args.command == "aggregate":
        aggregate(args.days_dir, args.prior, args.output)


if __name__ == "__main__":
    main()

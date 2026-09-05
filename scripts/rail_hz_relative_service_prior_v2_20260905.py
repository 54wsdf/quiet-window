from __future__ import annotations

import argparse
import json
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import scripts.rail_hz_daily_service_field_init_20260905 as base

# The Line-B branch mapping has already been topology-audited in the MPPD repository.
# Only these audited names are hard-bound here. Absolute timetable timestamps are never
# promoted to 2019 realized truth; only relative arrival-to-arrival lags are reused.
B_BRANCH_NAME_TO_NODE = {
    "客运中心": 20,
    "乔司南": 28,
    "乔司": 29,
    "翁梅": 30,
    "余杭高铁站": 31,
    "南苑": 32,
    "临平": 33,
}


def _read_l1_partial_branch_lags(timetable: Path) -> dict[tuple[str, int, int], list[float]]:
    wb = load_workbook(timetable, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    columns = [str(x) if x is not None else "" for x in next(rows)]
    idx = {name: i for i, name in enumerate(columns)}
    required = ["Train ID", "Line", "Direction", "Station Order", "Station Name", "Arrival Time"]
    missing = [x for x in required if x not in idx]
    if missing:
        raise SystemExit(f"missing timetable columns: {missing}")

    trains: dict[tuple[str, str], list[tuple[int, str, Any]]] = defaultdict(list)
    for row in rows:
        if str(row[idx["Line"]]) != "L1":
            continue
        train = str(row[idx["Train ID"]])
        direction = str(row[idx["Direction"]])
        order = int(row[idx["Station Order"]])
        station = str(row[idx["Station Name"]])
        arr = row[idx["Arrival Time"]]
        trains[(train, direction)].append((order, station, arr))
    wb.close()

    out: dict[tuple[str, int, int], list[float]] = defaultdict(list)
    for (_train, direction), stops in trains.items():
        stops = sorted(stops, key=lambda x: x[0])
        for a, b in zip(stops, stops[1:]):
            ua = B_BRANCH_NAME_TO_NODE.get(a[1])
            vb = B_BRANCH_NAME_TO_NODE.get(b[1])
            if ua is None or vb is None:
                continue
            if a[2] is None or b[2] is None:
                continue
            lag = (b[2] - a[2]).total_seconds()
            if 0 < lag <= base.MAX_EDGE_LAG_S:
                out[(direction, ua, vb)].append(float(lag))
    return out


def _median(vals: list[float]) -> float | None:
    return statistics.median(vals) if vals else None


def _quantile(vals: list[float], p: float) -> float | None:
    return base.quantile(vals, p) if vals else None


def build(timetable: Path, output: Path) -> dict[str, Any]:
    tmp = output.with_suffix(".base.json")
    prior = base.parse_timetable_priors(timetable, tmp)
    tmp.unlink(missing_ok=True)

    partial = _read_l1_partial_branch_lags(timetable)
    edge_lookup = {
        (str(e["path_id"]), str(e["direction"]), int(e["from_station"]), int(e["to_station"])): e
        for e in prior["edges"]
    }

    # Preserve every directly observed structural edge and annotate its evidence class.
    for e in prior["edges"]:
        if e.get("arrival_to_arrival_lag_sec_median") is not None:
            e["prior_source_class"] = "AUX_FULL_ROUTE_DIRECT_RELATIVE_LAG"
            e["prior_support_n"] = int(e.get("n", 0))
            e["prior_is_fallback"] = False

    # 1) Recover B_main Up by the opposite-direction physical edge when the auxiliary
    # workbook does not contain a full-length Up train. This is a structural prior only.
    bmain_up = list(reversed(base.LINE_PATHS["B_main"]["nodes"]))
    for u, v in zip(bmain_up, bmain_up[1:]):
        e = edge_lookup[("B_main", "Up", u, v)]
        if e.get("arrival_to_arrival_lag_sec_median") is not None:
            continue
        rev = edge_lookup.get(("B_main", "Down", v, u))
        if rev and rev.get("arrival_to_arrival_lag_sec_median") is not None:
            for k in ("arrival_to_arrival_lag_sec_median", "arrival_to_arrival_lag_sec_p10", "arrival_to_arrival_lag_sec_p90"):
                e[k] = rev[k]
            e["n"] = 0
            e["prior_source_class"] = "OPPOSITE_DIRECTION_PHYSICAL_EDGE_STRUCTURAL_FALLBACK"
            e["prior_support_n"] = int(rev.get("n", 0))
            e["prior_is_fallback"] = True

    # 2) B_branch shares nodes 0..20 with B_main. Reuse the same physical-edge prior
    # instead of requiring a fictitious full-length branch train in the timetable.
    for direction in ("Down", "Up"):
        nodes = list(base.LINE_PATHS["B_branch"]["nodes"])
        if direction == "Up":
            nodes = list(reversed(nodes))
        for u, v in zip(nodes, nodes[1:]):
            e = edge_lookup[("B_branch", direction, u, v)]
            if e.get("arrival_to_arrival_lag_sec_median") is not None:
                continue
            if u <= 20 and v <= 20:
                src = edge_lookup.get(("B_main", direction, u, v))
                if src and src.get("arrival_to_arrival_lag_sec_median") is not None:
                    for k in ("arrival_to_arrival_lag_sec_median", "arrival_to_arrival_lag_sec_p10", "arrival_to_arrival_lag_sec_p90"):
                        e[k] = src[k]
                    e["n"] = 0
                    e["prior_source_class"] = "SHARED_TRACK_B_MAIN_STRUCTURAL_REUSE"
                    e["prior_support_n"] = int(src.get("prior_support_n", src.get("n", 0)))
                    e["prior_is_fallback"] = True
                    continue

            vals = partial.get((direction, u, v), [])
            if vals:
                e["arrival_to_arrival_lag_sec_median"] = _median(vals)
                e["arrival_to_arrival_lag_sec_p10"] = _quantile(vals, 0.1)
                e["arrival_to_arrival_lag_sec_p90"] = _quantile(vals, 0.9)
                e["n"] = len(vals)
                e["prior_source_class"] = "AUX_PARTIAL_TRAIN_DIRECT_RELATIVE_LAG"
                e["prior_support_n"] = len(vals)
                e["prior_is_fallback"] = False
                continue

            rev_vals = partial.get(("Up" if direction == "Down" else "Down", v, u), [])
            if rev_vals:
                e["arrival_to_arrival_lag_sec_median"] = _median(rev_vals)
                e["arrival_to_arrival_lag_sec_p10"] = _quantile(rev_vals, 0.1)
                e["arrival_to_arrival_lag_sec_p90"] = _quantile(rev_vals, 0.9)
                e["n"] = 0
                e["prior_source_class"] = "AUX_PARTIAL_TRAIN_OPPOSITE_DIRECTION_FALLBACK"
                e["prior_support_n"] = len(rev_vals)
                e["prior_is_fallback"] = True

    # 3) Last-resort safety net: never delete a physical edge merely because the
    # auxiliary schedule is sparse. Use a line-specific median as an explicit weak
    # structural prior with uncertainty; downstream AFC evidence is allowed to move it.
    direct_by_aux: dict[str, list[float]] = defaultdict(list)
    for e in prior["edges"]:
        val = e.get("arrival_to_arrival_lag_sec_median")
        if val is not None and not e.get("prior_is_fallback", False):
            direct_by_aux[str(e["aux_line"])].append(float(val))
    global_vals = [v for vals in direct_by_aux.values() for v in vals]
    global_median = statistics.median(global_vals) if global_vals else 180.0

    for e in prior["edges"]:
        if e.get("arrival_to_arrival_lag_sec_median") is not None:
            continue
        vals = direct_by_aux.get(str(e["aux_line"]), [])
        val = statistics.median(vals) if vals else global_median
        e["arrival_to_arrival_lag_sec_median"] = float(val)
        e["arrival_to_arrival_lag_sec_p10"] = None
        e["arrival_to_arrival_lag_sec_p90"] = None
        e["n"] = 0
        e["prior_source_class"] = "LINE_MEDIAN_WEAK_STRUCTURAL_FALLBACK"
        e["prior_support_n"] = len(vals)
        e["prior_is_fallback"] = True

    classes: dict[str, int] = defaultdict(int)
    for e in prior["edges"]:
        classes[str(e.get("prior_source_class", "UNCLASSIFIED"))] += 1

    gates = {
        "all_directed_physical_edges_have_finite_prior": all(e.get("arrival_to_arrival_lag_sec_median") is not None for e in prior["edges"]),
        "absolute_auxiliary_timestamp_not_promoted": prior["absolute_timestamp_policy"] == "FORBIDDEN_AS_2019_REALIZED_TRUTH",
        "missing_full_route_sequence_is_not_a_fatal_gate": True,
        "fallback_edges_are_explicitly_labeled": all((not e.get("prior_is_fallback", False)) or "FALLBACK" in str(e.get("prior_source_class", "")) or "REUSE" in str(e.get("prior_source_class", "")) for e in prior["edges"]),
    }
    prior["schema"] = "rail.hz-relative-service-prior.v2-uncertainty-aware-partial-route"
    prior["prior_policy"] = "FULL_ROUTE_SEQUENCES_ARE_HELPFUL_BUT_NOT_REQUIRED; PARTIAL_TRAINS_SHARED_TRACK_AND_EXPLICIT_WEAK_FALLBACKS_RETAIN_PHYSICAL_EDGES"
    prior["prior_source_class_counts"] = dict(sorted(classes.items()))
    prior["integrity_gates"] = gates
    prior["status"] = "QUALIFIED_UNCERTAINTY_AWARE_RELATIVE_SERVICE_PRIOR" if all(gates.values()) else "RELATIVE_SERVICE_PRIOR_GATE_FAILED"
    output.write_text(json.dumps(prior, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"status": prior["status"], "source_classes": prior["prior_source_class_counts"], "integrity_gates": gates}, ensure_ascii=False, indent=2))
    if not all(gates.values()):
        raise SystemExit(2)
    return prior


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--timetable", type=Path, required=True)
    p.add_argument("--output", type=Path, required=True)
    a = p.parse_args()
    build(a.timetable, a.output)


if __name__ == "__main__":
    main()
